"""
Export action for auto-routes.
"""
from typing import Callable, Coroutine, Any
import io
import csv
from datetime import date, datetime
from enum import Enum
import html2text
from edgy import QuerySet

from fastapi import APIRouter, Query, HTTPException, Response
from starlette.responses import StreamingResponse

from core.api_route_model.actions import BaseApiRouteAction
from core.api_route_model.params import inject_order_by, OrderByQuery, FieldSelectorHeader, \
    optimize_query_filter_fields, FilterHeader, filter_query, InvalidFilterError
from core.api_route_model.params.field_selector import clean_field_names_from_input
from core.metadata_model.utils import get_field_label_from_path

from core.api_route_model.registry import TypeModel, RouteModelActionOptions
from core.orm.utils import get_value_from_path


class ExportApiRouteAction(BaseApiRouteAction):
    """Action for exporting model instances."""

    name = "export"

    @classmethod
    def register_route(
        cls,
        router: APIRouter,
        model_cls: TypeModel,
        options: RouteModelActionOptions
    ) -> None:
        """Register the export route."""
        router.add_api_route(**{
            "path": "/export",
            "endpoint": generate_export_items(model_cls),
            "methods": ["GET"],
            "summary": f"Export {model_cls.__name__} items",
            "description": f"Retrieve a export of {model_cls.__name__} items",
            **options,
        })


def generate_export_items[M = TypeModel](model_cls: type[M]) -> Callable[[str, int, int, str, str, str, int], Coroutine[Any, Any, Response]]:
    async def export_items(
            format: str = Query('csv', description="Export format (csv, xlsx, ods)"),
            limit: int | None = Query(None),
            offset: int = Query(0, ge=0),
            order_by: str | None = OrderByQuery(),
            fields: str | None = FieldSelectorHeader(),
            filters: str | None = FilterHeader(),
            list_id: int | None = Query(None, description="Filter by list ID"),
    ) -> Response:
        query = model_cls.query

        if list_id:
            query = query.filter(lists__id=list_id)

        return await export_items_action(
            model_cls,
            format,
            query,
            limit=limit,
            offset=offset,
            order_by=order_by,
            fields=fields,
            filters=filters,
        )

    return export_items


async def export_items_action[M = TypeModel](
    model_cls: type[M],
    format: str,
    query: QuerySet | None = None,
    limit: int | None = None,
    offset: int = 0,
    order_by: str | None = None,
    fields: str | None = None,
    filters: str | None = None,
) -> Response:
    try:
        query = query or model_cls.query
        query = filter_query(query, filters)
        query = inject_order_by(query, order_by)
        query = optimize_query_filter_fields(query, fields)

        items = await query.limit(limit).offset(offset).all()
    except InvalidFilterError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception as e:
        if filters:
            raise HTTPException(status_code=422, detail='Invalid filters')
        else:
            raise e

    field_names = clean_field_names_from_input(model_cls, fields)

    if not field_names or len(field_names) == 1 and field_names[0] == 'id':
        field_names = [field_name for field_name, field in model_cls.meta.fields.items()
                       if not field.exclude and not hasattr(field, 'target')]

    data_rows = []

    for item in items:
        row_data = []

        for field_name in field_names:
            row_data.append(format_value(await get_value_from_path(item, field_name)))

        data_rows.append(row_data)

    field_labels = [get_field_label_from_path(model_cls, field_name) for field_name in field_names]

    if format.lower() == 'csv':
        return generate_csv_export(field_labels, data_rows, f"{model_cls.__name__}_export.csv")

    if format.lower() == 'xlsx':
        return generate_xlsx_export(field_labels, data_rows, f"{model_cls.__name__}_export.xlsx")

    if format.lower() == 'ods':
        return generate_ods_export(field_labels, data_rows, f"{model_cls.__name__}_export.ods")

    raise HTTPException(status_code=400, detail=f"Unsupported export format: {format}")


def format_value(value: Any) -> str | None:
    """Format a value for export based on its type."""
    if value is None:
        return None
    elif isinstance(value, date):
        return value.strftime('%d/%m/%Y')
    elif isinstance(value, datetime):
        return value.strftime('%d/%m/%Y %H:%M:%S')
    elif isinstance(value, bool):
        return str(value).lower()
    elif isinstance(value, (int, float)):
        return str(value)
    elif isinstance(value, Enum):
        return value.value
    else:
        # Clean HTML tags from text values using html2text
        text_value = str(value)
        if '<' in text_value and '>' in text_value:
            h = html2text.HTML2Text()
            h.ignore_links = True
            h.ignore_images = True
            return h.handle(text_value).strip()
        return text_value


def generate_csv_export(field_names: list[str], data_rows: list[list[str]], filename: str) -> Response:
    """Generate a CSV export file."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(field_names)
    writer.writerows(data_rows)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def generate_xlsx_export(field_names: list[str], data_rows: list[list[str]], filename: str) -> Response:
    """Generate an XLSX export file."""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl library is not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Export"

    # Write header
    for col_idx, field in enumerate(field_names, 1):
        ws.cell(row=1, column=col_idx, value=field)

    # Write data
    for row_idx, row_data in enumerate(data_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column width
    for col_idx, field in enumerate(field_names, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(field), 15)

    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def generate_ods_export(field_names: list[str], data_rows: list[list[str]], filename: str) -> Response:
    """Generate an ODS export file."""
    try:
        from pyexcel_ods3 import save_data
    except ImportError:
        raise HTTPException(status_code=500, detail="pyexcel-ods3 library is not installed")

    # Prepare data for pyexcel-ods3
    sheet_data = [field_names]  # Header row
    
    # Convert None values to empty strings to avoid KeyError with pyexcel_ods3
    processed_rows = []
    for row in data_rows:
        processed_row = []
        for value in row:
            processed_row.append("" if value is None else value)
        processed_rows.append(processed_row)
    
    sheet_data.extend(processed_rows)  # Add processed data rows

    # Create ODS file
    data = {"Export": sheet_data}
    output = io.BytesIO()
    save_data(output, data)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.oasis.opendocument.spreadsheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
